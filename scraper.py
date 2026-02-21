"""
Supports loading existing Excel data and merging/deduplicating.
"""

import re
import time
from datetime import datetime
from urllib.parse import quote
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCROLL_PAUSE = 1.5
DETAIL_PAUSE = 1.0
MAX_SCROLLS = 15


# ──────────────────────────────────────────────────────
# EXISTING FILE LOADER
# ──────────────────────────────────────────────────────

def make_dedup_key(name, address):
    """Create a normalised key for deduplication."""
    name = re.sub(r'[^a-z0-9]', '', (name or '').lower())
    address = re.sub(r'[^a-z0-9]', '', (address or '').lower())
    return f"{name}|{address}"


def make_dedup_key_from_store(store):
    return make_dedup_key(store.get("name", ""), store.get("address", ""))


def load_existing_excel(filepath):
    """
    Load stores from an existing Excel file.
    Tries to auto-detect columns by header names (case-insensitive).
    Returns (list_of_store_dicts, set_of_dedup_keys).
    """
    stores = []
    seen_keys = set()

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return [], set(), f"Could not open file: {e}"

    # Try the first sheet, or one named with "store" in it
    ws = None
    for name in wb.sheetnames:
        if "store" in name.lower() or "all" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Read headers from row 1 or row 2 (some files have a title row)
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        wb.close()
        return [], set(), "File is empty"

    # Find the header row (look for a row containing "name" or "store name")
    header_row_idx = None
    headers = []
    for idx, row in enumerate(rows[:5]):  # Check first 5 rows
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        if any("name" in h for h in row_lower):
            header_row_idx = idx
            headers = row_lower
            break

    if header_row_idx is None:
        wb.close()
        return [], set(), "Could not find header row (looking for a column containing 'name')"

    # Map column names to indices
    col_map = {}
    name_aliases = {
        "name": ["store name", "name", "shop name", "business name", "store"],
        "address": ["address", "formatted address", "location", "full address"],
        "phone": ["phone", "phone number", "telephone", "tel", "phone no", "contact"],
        "rating": ["rating", "stars", "google rating"],
        "total_reviews": ["reviews", "total reviews", "review count", "no of reviews", "ratings count"],
        "category": ["category", "type", "primary type", "business type", "store type"],
        "website": ["website", "web", "url", "site", "website url"],
        "opening_hours": ["opening hours", "hours", "open hours", "timings"],
        "postcode": ["postcode", "postcode area", "post code", "zip", "outcode"],
        "latitude": ["latitude", "lat"],
        "longitude": ["longitude", "lng", "lon", "long"],
        "google_maps_url": ["google maps url", "google maps", "maps url", "maps link", "google maps link"],
    }

    for field, aliases in name_aliases.items():
        for i, h in enumerate(headers):
            if any(alias == h or alias in h for alias in aliases):
                col_map[field] = i
                break

    if "name" not in col_map:
        wb.close()
        return [], set(), "Could not find a 'Store Name' or 'Name' column"

    # Parse data rows
    data_rows = rows[header_row_idx + 1:]
    for row in data_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def get_val(field, default="N/A"):
            idx = col_map.get(field)
            if idx is not None and idx < len(row) and row[idx] is not None:
                val = str(row[idx]).strip()
                return val if val else default
            return default

        def get_num(field, default="N/A"):
            idx = col_map.get(field)
            if idx is not None and idx < len(row) and row[idx] is not None:
                try:
                    return float(row[idx])
                except (ValueError, TypeError):
                    return default
            return default

        def get_int(field, default=0):
            idx = col_map.get(field)
            if idx is not None and idx < len(row) and row[idx] is not None:
                try:
                    return int(float(row[idx]))
                except (ValueError, TypeError):
                    return default
            return default

        name_val = get_val("name", "")
        if not name_val or name_val == "N/A":
            continue

        store = {
            "name": name_val,
            "address": get_val("address"),
            "phone": get_val("phone"),
            "rating": get_num("rating"),
            "total_reviews": get_int("total_reviews"),
            "category": get_val("category"),
            "website": get_val("website"),
            "opening_hours": get_val("opening_hours"),
            "postcode": get_val("postcode", "—"),
            "latitude": get_val("latitude", ""),
            "longitude": get_val("longitude", ""),
            "google_maps_url": get_val("google_maps_url"),
            "_source": "existing",
        }

        key = make_dedup_key_from_store(store)
        if key not in seen_keys and key != "|":
            seen_keys.add(key)
            stores.append(store)

    wb.close()
    return stores, seen_keys, None


# ──────────────────────────────────────────────────────
# GOOGLE MAPS SCRAPING
# ──────────────────────────────────────────────────────

def build_url(query, postcode, location):
    full_query = f"{query} near {postcode} {location}"
    return f"https://www.google.com/maps/search/{quote(full_query)}"


def accept_cookies(page):
    for btn_text in ["Accept all", "Reject all"]:
        try:
            btn = page.locator(f'button:has-text("{btn_text}")').first
            if btn.is_visible(timeout=2000):
                btn.click()
                page.wait_for_timeout(1000)
                return True
        except Exception:
            continue
    return False


def scroll_results(page):
    feed_selector = 'div[role="feed"]'
    try:
        page.wait_for_selector(feed_selector, timeout=8000)
    except PWTimeout:
        return

    prev_count = 0
    stale_rounds = 0

    for i in range(MAX_SCROLLS):
        page.evaluate(f'''
            const feed = document.querySelector('{feed_selector}');
            if (feed) feed.scrollTop = feed.scrollHeight;
        ''')
        page.wait_for_timeout(int(SCROLL_PAUSE * 1000))

        try:
            end_text = page.locator("text=You've reached the end of the list").first
            if end_text.is_visible(timeout=500):
                break
        except Exception:
            pass

        items = page.locator(f'{feed_selector} > div > div > a[href*="/maps/place/"]')
        current_count = items.count()

        if current_count == prev_count:
            stale_rounds += 1
            if stale_rounds >= 3:
                break
        else:
            stale_rounds = 0
        prev_count = current_count


def extract_listing_links(page):
    return page.evaluate('''
        () => {
            const anchors = document.querySelectorAll('a[href*="/maps/place/"]');
            const urls = new Set();
            for (const a of anchors) {
                if (a.href && a.href.includes('/maps/place/')) urls.add(a.href);
            }
            return [...urls];
        }
    ''')


def extract_place_details(page):
    store = {
        "name": "N/A", "address": "N/A", "phone": "N/A",
        "website": "N/A", "rating": "N/A", "total_reviews": 0,
        "category": "N/A", "opening_hours": "N/A",
        "latitude": "", "longitude": "", "google_maps_url": page.url,
    }

    try:
        name_el = page.locator('h1').first
        if name_el.is_visible(timeout=3000):
            store["name"] = name_el.inner_text().strip()
    except Exception:
        pass

    try:
        rating_el = page.locator('div[role="img"][aria-label*="stars"]').first
        if rating_el.is_visible(timeout=1000):
            label = rating_el.get_attribute("aria-label") or ""
            match = re.search(r'([\d.]+)\s*star', label)
            if match:
                store["rating"] = float(match.group(1))
    except Exception:
        pass

    try:
        review_el = page.locator('button[aria-label*="reviews"]').first
        if review_el.is_visible(timeout=1000):
            label = review_el.get_attribute("aria-label") or review_el.inner_text()
            match = re.search(r'([\d,]+)\s*review', label)
            if match:
                store["total_reviews"] = int(match.group(1).replace(",", ""))
    except Exception:
        pass

    try:
        cat_el = page.locator('button[jsaction*="category"]').first
        if cat_el.is_visible(timeout=1000):
            store["category"] = cat_el.inner_text().strip()
    except Exception:
        pass

    try:
        info_items = page.locator('button[data-item-id], a[data-item-id]')
        count = info_items.count()
        for idx in range(count):
            item = info_items.nth(idx)
            try:
                data_id = item.get_attribute("data-item-id") or ""
                aria = item.get_attribute("aria-label") or ""
                text = item.inner_text().strip() if not aria else aria

                if data_id.startswith("address"):
                    store["address"] = text.replace("Address: ", "")
                elif data_id.startswith("phone"):
                    store["phone"] = text.replace("Phone: ", "")
                elif data_id.startswith("authority"):
                    website = item.get_attribute("href") or text.replace("Website: ", "")
                    store["website"] = website
            except Exception:
                continue
    except Exception:
        pass

    if store["phone"] == "N/A":
        try:
            phone_btn = page.locator('button[aria-label*="Phone:"]').first
            if phone_btn.is_visible(timeout=500):
                label = phone_btn.get_attribute("aria-label") or ""
                store["phone"] = label.replace("Phone: ", "").strip()
        except Exception:
            pass

    try:
        hours_el = page.locator(
            'div[aria-label*="Monday"], div[aria-label*="Sunday"], '
            'button[aria-label*="hours"], div[aria-label*="hour"]'
        ).first
        if hours_el.is_visible(timeout=1000):
            aria = hours_el.get_attribute("aria-label") or ""
            if aria:
                hours_text = aria.replace("Hours ", "").replace(". Hide open hours for the week", "")
                days = re.split(r'(?=Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)', hours_text)
                days = [d.strip().rstrip(";., ") for d in days if d.strip()]
                if days:
                    store["opening_hours"] = "\n".join(days)
    except Exception:
        pass

    try:
        coord_match = re.search(r'@(-?[\d.]+),(-?[\d.]+)', page.url)
        if coord_match:
            store["latitude"] = float(coord_match.group(1))
            store["longitude"] = float(coord_match.group(2))
    except Exception:
        pass

    return store


def scrape_postcode(page, query, postcode, location):
    url = build_url(query, postcode, location)
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2500)
    except Exception:
        return []

    scroll_results(page)
    links = extract_listing_links(page)

    if not links:
        try:
            h1 = page.locator('h1').first
            if h1.is_visible(timeout=2000):
                name = h1.inner_text().strip()
                if name and "results" not in name.lower():
                    store = extract_place_details(page)
                    store["google_maps_url"] = page.url
                    return [store]
        except Exception:
            pass
        return []

    stores = []
    for link in links:
        try:
            page.goto(link, wait_until="domcontentloaded", timeout=15000)
            page.wait_for_timeout(int(DETAIL_PAUSE * 1000))
            store = extract_place_details(page)
            store["google_maps_url"] = link
            stores.append(store)
        except Exception:
            continue

    return stores


def scrape_postcodes(query, location, postcodes,
                     existing_keys=None,
                     progress_callback=None, store_callback=None):
    """
    Main entry point. Scrapes Google Maps for each postcode.
    existing_keys: set of dedup keys from an uploaded file to skip.
    Returns (new_results, postcode_summary).
    """
    if existing_keys is None:
        existing_keys = set()

    def log(msg):
        if progress_callback:
            progress_callback(msg)

    new_results = []
    seen_keys = set(existing_keys)  # Start with existing keys so we skip duplicates
    postcode_summary = {}

    with sync_playwright() as p:
        log("Launching browser...")
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-GB",
        )
        page = context.new_page()

        page.goto("https://www.google.com/maps", wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)
        accept_cookies(page)
        page.wait_for_timeout(1000)

        for pc_idx, postcode in enumerate(postcodes):
            log(f"[{pc_idx+1}/{len(postcodes)}] Scraping postcode: {postcode}")

            stores_in_pc = scrape_postcode(page, query, postcode, location)
            new_count = 0
            skipped = 0
            phone_count = 0
            ratings = []

            for store in stores_in_pc:
                key = make_dedup_key_from_store(store)
                if key in seen_keys:
                    skipped += 1
                    continue

                seen_keys.add(key)
                store["postcode"] = postcode
                store["_source"] = "new"
                new_results.append(store)
                new_count += 1

                if store["phone"] != "N/A":
                    phone_count += 1
                if store["rating"] != "N/A":
                    ratings.append(store["rating"])

                if store_callback:
                    store_callback(store)

            avg_r = sum(ratings) / len(ratings) if ratings else None
            postcode_summary[postcode] = {
                "count": new_count,
                "phone_count": phone_count,
                "avg_rating": avg_r,
            }

            skip_msg = f", {skipped} already in sheet" if skipped > 0 else ""
            log(
                f"[{pc_idx+1}/{len(postcodes)}] {postcode}: "
                f"{len(stores_in_pc)} found, {new_count} new{skip_msg} — "
                f"Total new: {len(new_results)}"
            )

            if pc_idx < len(postcodes) - 1:
                time.sleep(1)

        browser.close()

    return new_results, postcode_summary


# ──────────────────────────────────────────────────────
# EXCEL EXPORT (supports merging existing + new)
# ──────────────────────────────────────────────────────

def create_excel(all_results, postcode_summary, output_path, query,
                 existing_stores=None):
    """
    Create formatted Excel. If existing_stores is provided, they are
    included in the sheet above the new results, clearly labelled.
    """
    if existing_stores is None:
        existing_stores = []

    # Combine: existing first, then new
    combined = []
    for s in existing_stores:
        s["_source"] = s.get("_source", "existing")
        combined.append(s)
    for s in all_results:
        s["_source"] = s.get("_source", "new")
        combined.append(s)

    # Sort by postcode then name
    combined.sort(key=lambda x: (x.get("postcode", "ZZZ"), x.get("name", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "All Stores"

    title_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor="F2F7FC")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    pc_fill = PatternFill("solid", fgColor="D6E4F0")
    new_fill = PatternFill("solid", fgColor="E8F5E9")     # Light green for new stores
    existing_fill = PatternFill("solid", fgColor="FFF8E1") # Light yellow for existing

    columns = [
        ("#", 5, True), ("Source", 10, True), ("Postcode", 12, True),
        ("Store Name", 30, False), ("Address", 42, False),
        ("Phone Number", 20, True), ("Rating", 8, True),
        ("Reviews", 10, True), ("Category", 22, False), ("Website", 35, False),
        ("Opening Hours", 40, False), ("Latitude", 12, True),
        ("Longitude", 12, True), ("Google Maps URL", 40, False),
    ]
    num_cols = len(columns)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_text = f'"{query}" — {datetime.now().strftime("%d %b %Y %H:%M")}'
    if existing_stores:
        title_text += f'  |  {len(existing_stores)} existing + {len(all_results)} new'
    ws["A1"].value = title_text
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    for col_idx, (header, width, _) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data
    for i, store in enumerate(combined):
        row = i + 3
        source = store.get("_source", "new")
        source_label = "Existing" if source == "existing" else "★ New"

        values = [
            i + 1, source_label, store.get("postcode", "—"),
            store.get("name", "N/A"), store.get("address", "N/A"),
            store.get("phone", "N/A"), store.get("rating", "N/A"),
            store.get("total_reviews", 0), store.get("category", "N/A"),
            store.get("website", "N/A"), store.get("opening_hours", "N/A"),
            store.get("latitude", ""), store.get("longitude", ""),
            store.get("google_maps_url", "N/A"),
        ]

        row_fill = new_fill if source == "new" else existing_fill

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align if columns[col_idx - 1][2] else data_align

            # Source column styling
            if col_idx == 2:
                if source == "new":
                    cell.font = Font(name="Arial", bold=True, size=10, color="2E7D32")
                    cell.fill = new_fill
                else:
                    cell.font = Font(name="Arial", size=10, color="F57F17")
                    cell.fill = existing_fill
            elif col_idx == 3:
                cell.fill = pc_fill
                cell.font = Font(name="Arial", bold=True, size=10)
            else:
                if i % 2 == 1:
                    cell.fill = alt_fill

            if col_idx in (10, 14) and val and val != "N/A":
                cell.font = link_font
                try:
                    cell.hyperlink = val
                except Exception:
                    pass

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}{len(combined) + 2}"

    # ── Postcode Summary sheet ──
    ws2 = wb.create_sheet("Postcode Summary")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Results by Postcode (New stores only)"
    ws2["A1"].font = title_font
    ws2["A1"].fill = title_fill
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(["Postcode", "New Stores", "With Phone", "Avg Rating"], 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12

    green_fill = PatternFill("solid", fgColor="E2EFDA")
    zero_fill = PatternFill("solid", fgColor="FCE4EC")

    for i, (pc, info) in enumerate(sorted(postcode_summary.items())):
        row = i + 3
        ws2.cell(row=row, column=1, value=pc).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=row, column=2, value=info["count"]).font = data_font
        ws2.cell(row=row, column=3, value=info["phone_count"]).font = data_font
        ws2.cell(row=row, column=4,
                 value=round(info["avg_rating"], 1) if info["avg_rating"] else "—").font = data_font
        for col in range(1, 5):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = (green_fill if i % 2 == 0 else alt_fill) if info["count"] > 0 else zero_fill

    # ── Info sheet ──
    ws3 = wb.create_sheet("Scrape Info")
    info_data = [
        ("Search Query", query),
        ("Date Scraped", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Method", "Playwright browser automation (Free)"),
        ("", ""),
        ("Existing Stores Loaded", len(existing_stores)),
        ("New Stores Scraped", len(all_results)),
        ("Total Stores in File", len(combined)),
        ("Duplicates Skipped", "Auto-deduplicated by name + address"),
    ]
    for r, (label, val) in enumerate(info_data, 1):
        ws3.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=r, column=2, value=val).font = Font(name="Arial", size=10)
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 55

    wb.save(output_path)
